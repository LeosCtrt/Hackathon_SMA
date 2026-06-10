"""
Génère HDJ_Agent_modele_donnees.xlsx — contrat de données uniquement.

Ce fichier décrit les colonnes attendues par HDJ Agent.
Il ne contient aucune configuration métier (ressources, soignants, parcours, PMSI).
La configuration métier est dans core/config/hdj_metier.yaml.

Usage : python build_data_contract.py
"""
from __future__ import annotations

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─── Palette ──────────────────────────────────────────────────────────────────
C_HEADER_BG = "1A5276"   # bleu hospitalier foncé
C_HEADER_FG = "FFFFFF"
C_REQ_BG    = "D6EAF8"   # bleu clair = obligatoire
C_OPT_BG    = "EBF5FB"   # bleu très clair = optionnel
C_WARN_BG   = "FDEBD0"   # orange = pseudonymisation
C_TITLE_BG  = "2E4057"
C_LISEZMOI  = "154360"


def _header_font(bold=True, color=C_HEADER_FG) -> Font:
    return Font(name="Calibri", bold=bold, color=color, size=11)


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _border() -> Border:
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _set_col_widths(ws, widths: dict[str, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ─── Données du contrat ───────────────────────────────────────────────────────
COLUMNS = [
    {
        "nom_standard": "NUM SEJOUR",
        "alias_alternatifs": "NUM_SEJOUR, id_sejour, séjour_id",
        "role": "Identifiant unique du séjour hospitalier",
        "type_attendu": "Texte / Entier",
        "obligatoire": "OUI",
        "exemple": "2024-000123",
        "anonymisation": "Pseudonymiser si sensible — ne jamais exporter le numéro SIH brut",
        "validation": "Non nul, unique par ligne de séjour",
    },
    {
        "nom_standard": "NUM IPP PATIENT",
        "alias_alternatifs": "IPP, ipp_patient, patient_id",
        "role": "Identifiant permanent patient pseudonymisé (permet l'analyse de récurrence)",
        "type_attendu": "Texte / Entier",
        "obligatoire": "RECOMMANDÉ",
        "exemple": "IPP-00456",
        "anonymisation": "OBLIGATOIRE : pseudonymiser ou hasher avant export. Jamais de nom/prénom/date de naissance.",
        "validation": "Stable dans le temps pour un même patient",
    },
    {
        "nom_standard": "CODE DIAG",
        "alias_alternatifs": "code_diagnostic, cim10, diagnostic_principal",
        "role": "Code diagnostic principal CIM-10 (détermine l'éligibilité HDJ)",
        "type_attendu": "Texte (ex. E11, E10, E03)",
        "obligatoire": "OUI",
        "exemple": "E11",
        "anonymisation": "Non sensible au niveau code",
        "validation": "Format CIM-10 — lettre + 2 chiffres minimum",
    },
    {
        "nom_standard": "LISTE ACTES CCAM MVT",
        "alias_alternatifs": "actes_ccam, codes_ccam, liste_actes",
        "role": "Liste des actes CCAM du mouvement (séparés par ';'). Utilisé pour l'éligibilité HDJ.",
        "type_attendu": "Texte (codes séparés par ';')",
        "obligatoire": "RECOMMANDÉ",
        "exemple": "BLQP010;YYYY180",
        "anonymisation": "Non sensible",
        "validation": "Peut être vide — l'outil sous-estime alors les candidats HDJ (comportement volontaire)",
    },
    {
        "nom_standard": "TYPE SEJOUR",
        "alias_alternatifs": "type_sejour, sejour_type, typesej",
        "role": "Type administratif du séjour (EXT = consultation externe, HDJ = hôpital de jour)",
        "type_attendu": "Texte : EXT | HDJ | MCO | …",
        "obligatoire": "OUI",
        "exemple": "EXT",
        "anonymisation": "Non sensible",
        "validation": "Valeurs attendues : EXT (consultations), HDJ (déjà structurés). L'outil traite les deux.",
    },
    {
        "nom_standard": "DATE ENTREE SEJ",
        "alias_alternatifs": "date_entree, date_sejour, date_admission",
        "role": "Date d'entrée du séjour — permet le filtrage par période",
        "type_attendu": "Date (AAAA-MM-JJ ou JJ/MM/AAAA)",
        "obligatoire": "RECOMMANDÉ",
        "exemple": "2024-03-15",
        "anonymisation": "Conserver l'année et le mois — jour optionnel selon politique établissement",
        "validation": "Format date parseable. Plage cohérente avec la période analysée.",
    },
    {
        "nom_standard": "HEURE ENTREE SEJ",
        "alias_alternatifs": "heure_entree, time_in, h_entree",
        "role": "Heure d'arrivée du patient (analyse flux journalier)",
        "type_attendu": "Heure (HH:MM ou HH:MM:SS)",
        "obligatoire": "OPTIONNEL",
        "exemple": "08:30",
        "anonymisation": "Non sensible",
        "validation": "Format heure entre 00:00 et 23:59",
    },
    {
        "nom_standard": "HEURE SORTIE SEJ",
        "alias_alternatifs": "heure_sortie, time_out, h_sortie",
        "role": "Heure de sortie du patient (calcul durée de séjour)",
        "type_attendu": "Heure (HH:MM ou HH:MM:SS)",
        "obligatoire": "OPTIONNEL",
        "exemple": "16:45",
        "anonymisation": "Non sensible",
        "validation": "Doit être >= HEURE ENTREE SEJ",
    },
    {
        "nom_standard": "CODE ACTE",
        "alias_alternatifs": "acte_principal, code_acte_principal",
        "role": "Code acte CCAM principal (complète LISTE ACTES CCAM MVT)",
        "type_attendu": "Texte",
        "obligatoire": "OPTIONNEL",
        "exemple": "BLQP010",
        "anonymisation": "Non sensible",
        "validation": "Format CCAM 7–8 caractères",
    },
    {
        "nom_standard": "SPECIALITE OPERATEUR",
        "alias_alternatifs": "specialite, code_spe, spe_operateur",
        "role": "Spécialité médicale de l'opérateur (filtre endocrino-diabétologie)",
        "type_attendu": "Texte (code spécialité RPPS ou libellé)",
        "obligatoire": "OPTIONNEL",
        "exemple": "Endocrinologie",
        "anonymisation": "Non sensible",
        "validation": "Utilisé pour filtrer la spécialité analysée",
    },
]

MAPPING_EXAMPLE = [
    {"local": "IPP_PATIENT", "standard": "NUM IPP PATIENT", "transformation": "Hasher SHA-256 avant export"},
    {"local": "NUM_SEJ", "standard": "NUM SEJOUR", "transformation": "Pseudonymiser"},
    {"local": "DGP", "standard": "CODE DIAG", "transformation": "Aucune"},
    {"local": "ACTES_CCAM", "standard": "LISTE ACTES CCAM MVT", "transformation": "Séparer par ';' si champs multiples"},
    {"local": "TYP_SEJ", "standard": "TYPE SEJOUR", "transformation": "Mapper EXT/J/N vers EXT/HDJ/MCO"},
    {"local": "DAT_ENT", "standard": "DATE ENTREE SEJ", "transformation": "Formater AAAA-MM-JJ"},
    {"local": "H_ENT", "standard": "HEURE ENTREE SEJ", "transformation": "Formater HH:MM"},
    {"local": "H_SORT", "standard": "HEURE SORTIE SEJ", "transformation": "Formater HH:MM"},
]


# ─── Construction du classeur ─────────────────────────────────────────────────
def build_contract(output_path: str = "HDJ_Agent_modele_donnees.xlsx") -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _build_readme(wb)
    _build_colonnes(wb)
    _build_mapping(wb)

    wb.save(output_path)
    print(f"✓ {output_path} généré — contrat de données propre.")


def _build_readme(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("_LisezMoi")
    ws.sheet_properties.tabColor = C_LISEZMOI

    rows = [
        ("HDJ Agent — Contrat de données hospitalières", None, None),
        (None, None, None),
        ("Rôle de ce fichier", None, None),
        (
            "Ce fichier décrit les colonnes que HDJ Agent attend pour analyser l'activité hospitalière.",
            None, None,
        ),
        ("Il ne contient aucune configuration métier.", None, None),
        (None, None, None),
        ("Séparation stricte des responsabilités", None, None),
        ("Document", "Rôle", "Contenu"),
        (
            "HDJ_Agent_modele_donnees.xlsx (ce fichier)",
            "Contrat de données",
            "Colonnes attendues, formats, obligatoire/optionnel, exemples, anonymisation",
        ),
        (
            "core/config/hdj_metier.yaml",
            "Source de vérité métier",
            "Ressources HDJ, soignants, capacités, parcours, règles PMSI, contraintes",
        ),
        (
            "Données_Externes_*.xlsx",
            "Données hospitalières pseudonymisées",
            "Séjours réels anonymisés — exemple CHU Guyane endocrino-diabéto 2020–2026",
        ),
        (None, None, None),
        ("Règle d'architecture", None, None),
        ("Excel = données et colonnes attendues", None, None),
        ("YAML = organisation métier du service HDJ", None, None),
        ("Streamlit = visualisation et simulation (ne modifie jamais le YAML)", None, None),
        (None, None, None),
        ("Utilisation", None, None),
        (
            "1. Préparer un export pseudonymisé depuis le SIH selon les colonnes de la feuille 'Colonnes_Attendues'.",
            None, None,
        ),
        (
            "2. Renseigner le mapping de vos colonnes locales dans la feuille 'Mapping_Colonnes'.",
            None, None,
        ),
        (
            "3. Contacter l'équipe HDJ Agent pour intégrer votre YAML métier.",
            None, None,
        ),
        (None, None, None),
        ("Confidentialité et pseudonymisation", None, None),
        ("— Ne jamais inclure de nom, prénom, adresse ou date de naissance.", None, None),
        ("— L'IPP (identifiant patient permanent) doit être pseudonymisé ou hashé avant export.", None, None),
        ("— HDJ Agent n'expose jamais d'IPP brut dans ses sorties ou son interface.", None, None),
    ]

    for i, row in enumerate(rows, 1):
        for j, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=val)
            if i == 1:
                cell.font = Font(name="Calibri", bold=True, size=16, color=C_LISEZMOI)
            elif val in ("Séparation stricte des responsabilités", "Règle d'architecture",
                         "Utilisation", "Rôle de ce fichier", "Confidentialité et pseudonymisation"):
                cell.font = Font(name="Calibri", bold=True, size=12, color=C_HEADER_BG)
            elif i == 8 and j in (1, 2, 3):
                cell.font = Font(name="Calibri", bold=True, color="FFFFFF")
                cell.fill = _fill(C_HEADER_BG)

    _set_col_widths(ws, {"A": 55, "B": 25, "C": 55})


def _build_colonnes(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("Colonnes_Attendues")
    ws.sheet_properties.tabColor = "1A5276"

    headers = [
        "Nom colonne standard HDJ Agent",
        "Alias / variantes acceptées",
        "Rôle",
        "Type attendu",
        "Obligatoire",
        "Exemple",
        "Anonymisation / pseudonymisation",
        "Règles de validation",
    ]

    # En-tête
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=j, value=h)
        cell.font = _header_font()
        cell.fill = _fill(C_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border()
    ws.row_dimensions[1].height = 35

    # Données
    for i, col in enumerate(COLUMNS, 2):
        vals = [
            col["nom_standard"],
            col["alias_alternatifs"],
            col["role"],
            col["type_attendu"],
            col["obligatoire"],
            col["exemple"],
            col["anonymisation"],
            col["validation"],
        ]
        is_req = col["obligatoire"] == "OUI"
        is_warn = "OBLIGATOIRE" in col.get("anonymisation", "")
        bg = C_REQ_BG if is_req else C_OPT_BG

        for j, val in enumerate(vals, 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = _border()
            if j == 7 and is_warn:
                cell.fill = _fill(C_WARN_BG)
                cell.font = Font(name="Calibri", bold=True, size=10)
            else:
                cell.fill = _fill(bg)
        ws.row_dimensions[i].height = 50

    # Légende
    leg_row = len(COLUMNS) + 3
    ws.cell(row=leg_row, column=1, value="Légende").font = Font(bold=True)
    for j, (color, label) in enumerate([
        (C_REQ_BG, "OUI — colonne obligatoire"),
        (C_OPT_BG, "RECOMMANDÉ ou OPTIONNEL"),
        (C_WARN_BG, "Attention pseudonymisation requise"),
    ], 1):
        c = ws.cell(row=leg_row + 1, column=j, value=label)
        c.fill = _fill(color)
        c.border = _border()

    _set_col_widths(ws, {
        "A": 28, "B": 35, "C": 40, "D": 22,
        "E": 14, "F": 18, "G": 48, "H": 38,
    })
    ws.freeze_panes = "A2"


def _build_mapping(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("Mapping_Colonnes")
    ws.sheet_properties.tabColor = "28A745"

    ws.cell(row=1, column=1, value="Mapping de colonnes — exemple d'adaptation").font = Font(
        bold=True, size=13, color=C_HEADER_BG
    )
    ws.cell(row=2, column=1, value=(
        "Renseignez vos noms de colonnes locaux dans la colonne 'Nom colonne établissement'."
        " Les colonnes standard HDJ Agent sont fixes."
    )).font = Font(italic=True, size=10, color="555555")

    headers = [
        "Nom colonne établissement (local)",
        "Nom colonne standard HDJ Agent",
        "Transformation si nécessaire",
        "Validé",
    ]
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=j, value=h)
        cell.font = _header_font()
        cell.fill = _fill(C_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border()
    ws.row_dimensions[4].height = 30

    for i, row in enumerate(MAPPING_EXAMPLE, 5):
        vals = [row["local"], row["standard"], row["transformation"], "☐"]
        for j, val in enumerate(vals, 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.fill = _fill("F0F8FF")
            cell.border = _border()

    # Lignes vides pour l'établissement
    for i in range(len(MAPPING_EXAMPLE) + 5, len(MAPPING_EXAMPLE) + 10):
        for j in range(1, 5):
            ws.cell(row=i, column=j).border = _border()

    ws.cell(row=len(MAPPING_EXAMPLE) + 5, column=1, value="← Ajouter vos colonnes supplémentaires ici").font = Font(
        italic=True, color="888888"
    )

    _set_col_widths(ws, {"A": 35, "B": 32, "C": 45, "D": 10})
    ws.freeze_panes = "A5"


if __name__ == "__main__":
    build_contract()
